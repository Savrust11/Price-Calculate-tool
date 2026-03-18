"""
Phase 4 — Excel Output
=======================
Formats the final bid results into a clean, styled Excel workbook.
"""

import logging
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Colour scheme
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BID_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NO_BID_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NO_DATA_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PRIORITY_BID_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Column ordering for the output sheet
OUTPUT_COLUMNS = [
    "BoxNo",
    "BranchNo",
    "Brand",
    "Details",
    "Rank",
    "Listings Count",
    "Market Price",
    "Min Price",
    "Max Price",
    "Grade Multiplier",
    "Adjusted Price",
    "Max Bid",
    "BidAmount",
    "Priority Keyword",
    "Decision",
]

# Human-friendly header labels (Japanese)
HEADER_LABELS = {
    "BoxNo": "箱番号",
    "BranchNo": "枝番号",
    "Brand": "ブランド",
    "Details": "商品名",
    "Rank": "ランク",
    "Listings Count": "出品数",
    "Market Price": "市場価格 (¥)",
    "Min Price": "最低落札価格 (¥)",
    "Max Price": "最高落札価格 (¥)",
    "Grade Multiplier": "ランク係数",
    "Adjusted Price": "調整価格 (¥)",
    "Max Bid": "最大入札額 (¥)",
    "BidAmount": "最低入札額 (¥)",
    "Priority Keyword": "優先キーワード",
    "Decision": "判定",
}

# Columns to format as JPY integers
YEN_COLUMNS = {
    "Market Price", "Min Price", "Max Price",
    "Adjusted Price", "Max Bid", "BidAmount",
}


def write_output_excel(df: pd.DataFrame, output_path: str) -> Path:
    """Write the bid-decision DataFrame to a formatted Excel file.

    Parameters
    ----------
    df : DataFrame
        Output of bid_calculator.apply_bid_decisions().
    output_path : str
        Destination .xlsx path.

    Returns
    -------
    Path to the written file.
    """
    # Select and order columns (skip any that don't exist)
    cols = [c for c in OUTPUT_COLUMNS if c in df.columns]
    out = df[cols].copy()

    # Rename headers
    out.columns = [HEADER_LABELS.get(c, c) for c in cols]

    out_path = Path(output_path)
    out.to_excel(out_path, index=False, sheet_name="入札結果", engine="openpyxl")

    # --- Style the workbook ---
    from openpyxl import load_workbook

    wb = load_workbook(out_path)
    ws = wb["入札結果"]

    # Header row
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Find the Decision column index
    decision_col = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col_idx).value == "判定":
            decision_col = col_idx
            break

    # Data rows
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Colour-code the Decision cell
        if decision_col:
            decision_cell = ws.cell(row=row_idx, column=decision_col)
            val = str(decision_cell.value or "")
            if val == "入札":
                decision_cell.fill = BID_FILL
            elif val == "データ不足":
                decision_cell.fill = NO_DATA_FILL
            elif val == "優先入札":
                decision_cell.fill = PRIORITY_BID_FILL

    # Format yen columns as integer with comma separator
    yen_headers = {HEADER_LABELS.get(c, c) for c in YEN_COLUMNS}
    for col_idx in range(1, ws.max_column + 1):
        header_val = ws.cell(row=1, column=col_idx).value
        if header_val in yen_headers:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell.number_format = "#,##0"

    # Auto-fit column widths
    for col_idx in range(1, ws.max_column + 1):
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
        for row_idx in range(2, min(ws.max_row + 1, 52)):  # sample first 50 rows
            val = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(val or "")))
        adjusted = min(max_len + 4, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(out_path)
    logger.info("Output written to %s (%d rows)", out_path, ws.max_row - 1)
    return out_path
